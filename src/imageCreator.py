from __future__ import division
from operator import truediv
import sys
sys.path.append(os.environ['SHARED_PYTHON_LIBS'])

import dyTextUtils
reload(dyTextUtils)
from PySide import QtCore, QtGui
import json
from pprint import pprint
import os
import sys
import openpyxl
import MaxPlus

fname = os.path.join(os.path.dirname(os.path.abspath(__file__)), "imageCreator.ui")
formt, btype = MaxPlus.LoadUiType(fname)


def testValidValue(value):
	if value is not None:
		rawValue = r"%s"%value
		if not "none" in rawValue :
			return True
		else:
			return False


def noDigits(s):
	return ''.join([i for i in s if not i.isdigit()])


def convertToStr(strValue):
	if strValue != None:
		if isinstance(strValue, int) or  isinstance(strValue, float):
			strValue = str(strValue)
		elif ( isinstance(strValue, unicode)):
			dyTextUtils.fixText(strValue)
			#strValue = str(strValue.encode('utf8'))
		else:
			#strValue = strValue
			strValue = row[strValueCol].value
	return strValue


def getModelDicts(dict):
	modelDict = {}
	keylist = []
	for key in dict:
		if 'Model' in key:
			if not '_' in key:
				keylist.append( dyTextUtils.fixText(key) )

	for key in keylist:
		currentModelDict = {}
		for dictKey in dict:
			if key in dictKey:
				currentModelDict[ noDigits(dictKey) ] = dyTextUtils.fixText( dict[dictKey] )
		modelDict[key]= currentModelDict
	return modelDict


def getCategories(sheet):
	categorieslist = []
	for i in range(55):
		category = convertToStr( sheet.cell(row=1,column=i+1).value )
		if testValidValue(category):
			category = category.split(' ')[0]
			categorieslist.append(dyTextUtils.fixText (category) )
	return categorieslist


def getTasks(sheet,task_col,enviroment_col):
	tasklist =set()
	# collect tasks
	for row in sheet.iter_rows(min_row=2, min_col=1, max_row=400, max_col=400):
		if row[task_col].value is not None:
			tasklist.add( (convertToStr(row[task_col].value), convertToStr(row[enviroment_col].value)))
	return tasklist


def getConfigs(sheet,task,task_col):
	configlist = []
	for row in sheet.iter_rows(min_row=2, min_col=1, max_row=400, max_col=2):
		row_value = row[task_col].value
		row_value = convertToStr( row_value )
		if not row_value:
			continue
		if row_value != task:
			continue
		imageDict = {}
		row = row[task_col].row
		# find category and corrensponding value
		for i in range(55):
			val=convertToStr( sheet.cell(row=row,column=i+1).value )
			category = convertToStr( sheet.cell(row=1,column=i+1).value )
			if testValidValue(category):
					category =dyTextUtils.fixText (  category.split(' ')[0] )
			if testValidValue(val):
				imageDict[category] = dyTextUtils.fixText (val)
		configlist.append(imageDict)
	return configlist


def run(dict,maxscript,  renderRes="Full", saveRenderfile=True, skipExistingExrs=True,renderType="BackBurner"):

	try:
		fileinStr = 'fileIn("%s")'%maxscript
		MaxPlus.Core.EvalMAXScript(fileinStr)
		task = dict['Delprojekt']

		# reset max
		restStr = 'resetScene()'
		print restStr
		MaxPlus.Core.EvalMAXScript(restStr)

		# load enviroment
		loadEnviromentStr = 'loadEnviroment "%s"' %dict['Rum']
		print loadEnviromentStr
		MaxPlus.Core.EvalMAXScript(loadEnviromentStr)
		# set camera
		if dict.get('Camera')  is not None:
				camStr = 'viewport.setCamera $%s' %dict['Camera']
				print camStr
				MaxPlus.Core.EvalMAXScript(camStr)

		modelDict = getModelDicts(dict)
		sortedKeys=sorted(modelDict.keys(), key=lambda x:x.lower())

		for i in sortedKeys:
			model=modelDict[i]

			if model.get('Model_action')  is not None:
				actionStr = '%s()'%model['Model_action']
				print "\n### DO: ",  actionStr
				MaxPlus.Core.EvalMAXScript(actionStr)

			print "\n### DO: loadproduct",  model['Model']
			loadStr = 'loadproduct @"%s"'%model['Model']
			MaxPlus.Core.EvalMAXScript(loadStr)

			if model.get('Model_Elevation')  is not None:
				elevation = model['Model_Elevation']
			else:
				elevation = 0

			if model.get('Model_Dummy')  is not None:
				print "### DO: alignProduct ", model['Model_Dummy'] , elevation
				alignStr = 'alignProduct "%s" %s'%(model['Model_Dummy'] , elevation)
				MaxPlus.Core.EvalMAXScript(alignStr)

			if model.get('Model_Material')  is not None:
				if model.get('Model_OldMaterial')  is not None:
					newMaterial = model['Model_Material']
					oldMaterial = model['Model_OldMaterial']
					print "### DO: switchMaterial" , newMaterial ,oldMaterial
					matSwitchStr = 'switchMaterial "%s" "%s"'%(oldMaterial,newMaterial)
					MaxPlus.Core.EvalMAXScript(matSwitchStr)
		print "Render", task
		if "Artikel" in dict['Delprojekt']:
			rendername =  dict["Bild"]
		else:
			rendername =  dict['Delprojekt'].replace(".","")+"_"+dict["Bild"]

		renderStr = r'renderConfig "%s" "%s" renderRes:#%s saveRenderfile:%s  skipExistingExrs:%s renderType:#%s '%(
																						task,
																						rendername.replace(" ",""),
																						renderRes,
																						saveRenderfile,
																						skipExistingExrs,
																						renderType
																						)

		MaxPlus.Core.EvalMAXScript(renderStr)

	except Exception as e:
		print e.message, e.args
	print "\n"
	print "DONE"



class ImageCreatorWindow(btype, formt):
	def set_local_machineRender(self,endStr,use_local_machine=0):
		use_local_machineList = endStr.split('use_local_machine')[-1].split('\n')
		#use_local_machine = int(use_local_machineList[0])
		#use_local_machine = 1
		endStr = "%suse_local_machine %i\n%s"%(endStr.split('use_local_machine')[0],
						use_local_machine,
						use_local_machineList[-1]
						)
		return endStr


	def writeCFG(self,renderStr):
		localappData = os.getenv('LOCALAPPDATA')
		cfgPath = localappData+"/Autodesk/3dsMax/2017 - 64bit/ENU/en-US/plugcfg/"
		# filedir = 'c:/Users/jfo/Downloads/'
		file = open(cfgPath+'vray_dr.cfg', 'w')

		file.write(renderStr)
		file.close()


	def readCFG(self):
		# user =
		localappData = os.getenv('LOCALAPPDATA')
		cfgPath = localappData+"/Autodesk/3dsMax/2017 - 64bit/ENU/en-US/plugcfg/vray_dr.cfg"

		file = open(cfgPath, 'r')
		renderDict = {}

		l = file.readlines()
		endStr = ""

		for i in l :
			p = i.split(" ")
			if len(p) >= 3:
				key = p[0]
				on = int(p[1])
				port = int(p[2])
				renderDict[key] = {"on":on,"port":port}
			else:
				endStr = endStr + i
		return renderDict, endStr


	def setDrSlave(self,slave,status,renderDict):
		renderStr = ""
		for key in renderDict:
			if key == slave:
				str = "%s %i %i" %( key ,status,renderDict[key]["port"])
			else:
				str = "%s %i %i" %( key ,renderDict[key]["on"],renderDict[key]["port"])
			renderStr = renderStr + str + "\n"
		return renderStr


	def setDrSlaves(self):
		cfg = readCFG()
		renderDict = cfg[0]
		endStr = cfg[1]
		renderStr = setDrSlave("hp_z230_18", 0 ,renderDict )
		endStr = self.set_local_machineRender(endStr,use_local_machine=0)
		renderStr = renderStr + endStr
		writeCFG(renderStr)


	def populateDRslaves(self):
		self.drmodel.removeRows( 0, self.drmodel.rowCount() )
		cfg = self.readCFG()
		renderDict = cfg[0]
		endStr = cfg[1]
		use_local_machineList = endStr.split('use_local_machine')[-1].split('\n')
		use_local_machine = int(use_local_machineList[0])
		self.checkBox_useLocalMachine.setChecked(bool(use_local_machine))
		for slave in renderDict:
			# Create an item with a caption?
			item = QtGui.QStandardItem(slave)

			# Add a checkbox to it
			item.setCheckable(True)

			# Add the item to the model
			self.drmodel.appendRow(item)

			self.drlistView.setModel(self.drmodel)
			if self.checkBox_allDrSlaves.isChecked():
				item.setCheckState(QtCore.Qt.Checked)
			else:
				if renderDict[slave]['on'] is 1:
					item.setCheckState(QtCore.Qt.Checked)



	def commitDrConfigbtnPushed(self):
		cfg = self.readCFG()
		renderDict = cfg[0]
		endStr = cfg[1]
		for row in range(self.drmodel.rowCount()):
				item = self.drmodel.item(row)
				slave = item.text()
				if item.checkState() == QtCore.Qt.Checked:
					print slave
					renderDict[slave]['on'] = 1
					renderStr = self.setDrSlave(slave, 1 ,renderDict )
				else:
					renderDict[slave]['on'] = 0
					renderStr = self.setDrSlave(slave, 0 ,renderDict )

		if self.checkBox_useLocalMachine.isChecked():
			endStr = self.set_local_machineRender(endStr,use_local_machine=1)
		else:
			endStr = self.set_local_machineRender(endStr,use_local_machine=0)

		renderStr = renderStr + endStr
		print renderStr
		self.writeCFG(renderStr)


	def selectAllCheckDrSlaveChanged(self):
		model = self.drmodel
		for row in range(model.rowCount()):
			item = model.item(row)
			if item.isCheckable():
				if self.checkBox_allDrSlaves.isChecked():
					item.setCheckState(QtCore.Qt.Checked)
				else:
					item.setCheckState(QtCore.Qt.Unchecked)


	def populateTasks(self):
		self.tasks = getTasks(self.sheet,self.task_col,self.enviroment_col)
		tasks = list(self.tasks)
		tasks = [i[0] for i in tasks]
		tasks.sort(key=len, reverse=True) # sorts by descending length
		tasks.sort() # sorts normally by alphabetical order
		self.taskComboBox.clear()
		self.taskComboBox.addItems(tasks)


	def selectAllCheckChanged(self):
		model = self.ConfiglistView.model()
		for row in range(model.rowCount()):
			item = model.item(row)
			if item.isCheckable():
				if self.select_all_cb.isChecked():
					item.setCheckState(QtCore.Qt.Checked)
				else:
					item.setCheckState(QtCore.Qt.Unchecked)


	def selectFirstRow(self):
		model = self.ConfiglistView.model()
		item = model.item(0)
		item.setCheckState(QtCore.Qt.Checked)


	def populateConfigurations(self):
		self.model.removeRows( 0, self.model.rowCount() )
		for config in self.configList:
			# Create an item with a caption?
			item = QtGui.QStandardItem(config)

			# Add a checkbox to it
			item.setCheckable(True)

			# Add the item to the model
			self.model.appendRow(item)
			#item.setIcon(some_QIcon)
			self.ConfiglistView.setModel(self.model)
			if self.select_all_cb.isChecked():
				item.setCheckState(QtCore.Qt.Checked)


	def onInputFileButtonClicked(self):
		filename, filter = QtGui.QFileDialog.getOpenFileName(parent=self, caption='Open file', dir='.', filter='*.xlsx')
		if filename:
			self.orderPath.setText(filename)
			self.load()


	def runBtnClicked(self):
		model = self.ConfiglistView.model()
		skipExistingExrsBool = self.checkBox_skipExistingExrs.isChecked()
		renderRes = self.comboBox_renderRes.currentText()
		saveRenderfile = self.checkBox_saveBatchScene.isChecked()
		renderType =  self.comboBox_renderType.currentText()
		fileinStr = 'fileIn "%s"'%str(self.maxscript)
		MaxPlus.Core.EvalMAXScript(fileinStr)

		if self.allTasks_checkBox.isChecked():
			tasks = getTasks(self.sheet,self.task_col,self.enviroment_col)
			for curTask in tasks:
				curTaskStr = dyTextUtils.fixText(curTask[0])
				configdicts = getConfigs(self.sheet,curTaskStr,self.task_col)
				task =configdict['Delprojekt']
				for configdict in configdicts:
					run(configdict,self.maxscript,  renderRes=renderRes, saveRenderfile=saveRenderfile, skipExistingExrs=skipExistingExrsBool , renderType=renderType)
		else:
			for row in range(model.rowCount()):
				item = model.item(row)
				if item.checkState() != QtCore.Qt.Checked:
					continue
				configdict = getConfigs(self.sheet,self.task ,self.task_col)
				configdict = [i for i in configdict if i['Bild'] == item.text() ][0]
				run(configdict,self.maxscript,  renderRes=renderRes, saveRenderfile=saveRenderfile, skipExistingExrs=skipExistingExrsBool , renderType=renderType)


	def taskChanged(self, index):
		self.task  = self.taskComboBox.currentText()
		self.getConfigList()
		self.populateConfigurations()

	def initExcel(self):
		book = openpyxl.load_workbook(self.excelFile)
		self.sheet = book.get_sheet_by_name("Bilder")
		self.task_col = 1
		self.enviroment_col = 6
		ws = book.get_sheet_by_name("config")
		self.maxscript= ws.cell(row=1, column=2).value

	def getConfigList(self):
		self.configList = [i['Bild'] for  i in  getConfigs(self.sheet,self.task ,self.task_col)]


	def load(self):
		# init fields
		self.outputBaseDir = self.baseDir  + "Development/3D/renderoutput/"
		self.jsonFileDir = self.baseDir + "Development/3D/json/"
		self.jsonfile = self.jsonFileDir +"project_2.2.json"
		self.maxscript =  self.maxscriptdir  + "renderConfig.ms"
		self.excelFile = self.orderPath.text()
		self.task = '1.0'
		# init ui
		self.initExcel()
		self.getConfigList()
		self.populateTasks()
		self.populateConfigurations()
		self.populateDRslaves()


	def __init__(self, parent=None, baseDir='', maxscriptdir=''):
		self.baseDir = baseDir
		self.maxscriptdir = maxscriptdir
		btype.__init__(self)
		formt.__init__(self)
		self.setupUi(self)
		self.model = QtGui.QStandardItemModel()
		self.drmodel = QtGui.QStandardItemModel()
		# load fields
		self.load()

		# connect to slots
		self.runBtn.clicked.connect(self.runBtnClicked)
		self.taskComboBox.currentIndexChanged[str].connect(self.taskChanged)
		self.select_all_cb.stateChanged.connect(self.selectAllCheckChanged)
		self.checkBox_allDrSlaves.stateChanged.connect(self.selectAllCheckDrSlaveChanged)
		self.browse_pushButton.clicked.connect(self.onInputFileButtonClicked)
		self.commitDrConfigbtn.clicked.connect(self.commitDrConfigbtnPushed)



form = ImageCreatorWindow()
MaxPlus.AttachQWidgetToMax(form )
form.show()